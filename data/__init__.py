"""数据驱动工具 - 从 YAML / Excel 读取测试数据供 parametrize 使用。"""

from pathlib import Path

import yaml
from openpyxl import load_workbook

_data_dir = Path(__file__).parent


# ==================== YAML 数据驱动 ====================

def load_yaml_data(key: str, file_name: str = "test_data.yaml") -> list[dict]:
    """从 YAML 文件加载指定 key 的测试数据列表。

    Args:
        key: YAML 文件中的顶层 key，如 "login_phone"
        file_name: YAML 文件名，默认 test_data.yaml

    Returns:
        测试数据字典列表
    """
    with open(_data_dir / file_name, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return data.get(key, [])


def yaml_parametrize(key: str, file_name: str = "test_data.yaml"):
    """加载 YAML 测试数据并返回 (cases, ids) 元组，方便 parametrize 使用。"""
    cases = load_yaml_data(key, file_name)
    ids = [c.get("case_id", c.get("title", f"case_{i}")) for i, c in enumerate(cases)]
    return cases, ids


# ==================== Excel 数据驱动 ====================

def read_excel_data(file_name: str, sheet_name: str = None) -> list[dict]:
    """读取 Excel 文件，返回字典列表（每行一条数据，表头为 key）。

    Args:
        file_name: 文件名（相对于 data/ 目录）或绝对路径
        sheet_name: 工作表名，默认读取第一个 sheet

    Returns:
        [{"col1": val1, "col2": val2}, ...]
    """
    file_path = _data_dir / file_name
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h).strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        record = dict(zip(headers, row))
        if any(v is not None for v in row):
            data.append(record)

    wb.close()
    return data


def read_excel_parametrize(file_name: str, sheet_name: str = None):
    """读取 Excel 并返回 pytest.mark.parametrize 需要的格式。"""
    data = read_excel_data(file_name, sheet_name)
    if not data:
        return [], [], []

    argnames = list(data[0].keys())
    argvalues = [tuple(d.values()) for d in data]
    ids = [d.get("case_name", f"case_{i}") for i, d in enumerate(data)]
    return ids, argnames, argvalues
